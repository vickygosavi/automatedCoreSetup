import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog,ttk
#import tkinter
#from typing_extensions import ParamSpecKwargs
import selenium.webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import select
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import *
import os
import sys
import xlwings as xw
import pandas as pd

#from Run_Macro import Load_excel_data



'''
Interface - Tkinter
Automation - Selenium
Excel Macro run - xlwings
Excel read - Openpyxl

Variables - CSV location and Designation's dropdown options
If any thing in Designation page (client's page + /import/asyncImports/type/designation) is changed then change the 
respective string values of these variables.

Functions are created on the basis of requirements ; run_excel_macro and Openpyxl's code

Functions named, relative_path1 and relative_path_logo are created as Chromedriver.exe and logo can be part of the executable file.

Functions mentioned below above mentioned functions (including) 'on_open' are part of the selenium code which are then added as a 'command' in the respective button

TKinter -

An If - else logic to display all the input windows including GC city, GC state etc; if user select 'yes' to change/update the GC profile.
If user decides not to update GC profle then only Web address, Login ID and Password inputs are visible

Macro -
Select file function to access macro file
Run_exce_macro function to run the macro from selected file.

Grid property is used to arrange all the labels and buttons

'''


# ------------ Variables which are uploaded in Designation's dropdown ------------------

Dept = "Department code"
Desg = "Designations - Alias"
Desg_code = "Designations - Alias Code"
Numb_of_position = "Number Of Position"
Not_period = "count notice period employee"
Funct_area = "Functional Area"

# ------------------------------------ CSV upload ---------------------------------------

Business_unit_up = "D:/Import Files/Business Unit.csv"
Division_up = "D:/Import Files/Division.csv"
Designation_loc = "D:/Import Files/Designation Location.csv"
Dept_master = "D:/Import Files/Department Master.csv"
Dept_mapping = "D:/Import Files/Department Mapping.csv"
Cost_center = "D:/Import Files/Cost Center.csv"

Loc_type = "D:/Import Files/Location Type.csv"
City_type = "D:/Import Files/City Type.csv"
Locations_upload = "D:/Import Files/Location.csv"
Func_area = "D:/Import Files/Functional Area.csv"
Func_area_mapping = "D:/Import Files/Functional Area Mapping.csv"
#Band_upload = "D:/Import Files/Band.csv"
Grade_upload = "D:/Import Files/Grade.csv"
JobLevel_upload = "D:/Import Files/Joblevel.csv"
Contrib_level = "D:/Import Files/Contribution Level.csv"
Designation_name = "D:/Import Files/Designation Name.csv"
Designation = "D:/Import Files/Designation.csv"
Grade_in_designation = "D:/Import Files/Grade in Designation.csv"

# --------------------------------------- functions -------------------------------------
# --------------------------------- xlwings to run macro -------------------------------

# To open a dialog box and store selected file

def select_file():
    filetypes = (
        ('text files', '*.txt'),
        ('All files', '*.*')
    )

    global filename1
    #filename1 = StringVar()
    filename1 = filedialog.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)

# --------------------- access macro - Error check & Generate error --------------------

def run_excel_macro_validation():
    
    '''To run validation macro'''
        
    try:
        xl_app = xw.App(visible=False, add_book=False)
        wb = xl_app.books.open(filename1)

        #To check errors
        run_macro1 = wb.app.macro('CheckErrors.CheckforErrors') 

        #To generate Err report
        run_macro2 = wb.app.macro('Error_Report.Error_Report') 

        run_macro1()
        run_macro2()

        wb.save()
        wb.close()

        xl_app.quit()

    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)
        


# --------------------- access macro - Errors resolved ------------------------------

def run_validation_Allresolved_macro():
    
    '''To run validation macro'''
        
    try:
        xl_app = xw.App(visible=False, add_book=False)
        wb = xl_app.books.open(filename1)

        # To resolve errors
        run_macro3 = wb.app.macro('CheckErrors.ResolvedAllErrors')

        run_macro3()

        wb.save()
        wb.close()

        xl_app.quit()

    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)


# --------------------- access macro - To create core exports --------------------

def run_excel_macro():
    """
    Execute an Excel macro
    :param file_path: path to the Excel file holding the macro
    :return: None
    """
        
    try:
        xl_app = xw.App(visible=False, add_book=False)
        wb = xl_app.books.open(filename1)

        run_macro = wb.app.macro('New_Super_macro_v2.New_Super_macro_V2')
        run_macro()

        wb.save()
        wb.close()

        xl_app.quit()

    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)

# ------------------------------------- Tree view for Error report -------------------------------------------

def EntireTree():

    global root2
    root2 = tk.Tk()

    root2.geometry("850x650") # set the root2 dimensions

    #Frame for TreeView
    global frame1
    frame1 = tk.LabelFrame(root2, text="Excel Data")
    frame1.place(height=500, width=800,rely=0, relx=0.02)


    button2 = tk.Button(root2, text="Load File", command=lambda: Load_excel_data())
    button2.place(rely=0.80, relx=0.20)

    # The file/file path text
    global label_file
    label_file = ttk.Label(root2, text="No File Selected")
    label_file.place(rely=0.90, relx=0.45)


    ## Treeview Widget
    global tv1
    tv1 = ttk.Treeview(frame1)
    tv1.place(relheight=1, relwidth=1) # set the height and width of the widget to 100% of its container (frame1).

    treescrolly = tk.Scrollbar(frame1, orient="vertical", command=tv1.yview) # command means update the yaxis view of the widget
    treescrollx = tk.Scrollbar(frame1, orient="horizontal", command=tv1.xview) # command means update the xaxis view of the widget
    tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set) # assign the scrollbars to the Treeview Widget
    treescrollx.pack(side="bottom", fill="x") # make the scrollbar fill the x axis of the Treeview widget
    treescrolly.pack(side="right", fill="y") # make the scrollbar fill the y axis of the Treeview widget

    root2.mainloop()

# -------------------------- Functions which are being used inside EntireTree function -----------------------------------

def Load_excel_data():
    """If the file selected is valid this will load the file into the Treeview"""
    #file_path = label_file["text"]
    try:
        excel_filename = r"{}".format(filename1)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename,sheet_name='Error_Report')
        

    except ValueError:
        tk.messagebox.showerror("Information", "The file you have chosen is invalid")
        return None
    except FileNotFoundError:
        tk.messagebox.showerror("Information", f"No such file as {filename1}")
        return None

    clear_data()
    tv1["column"] = list(df.columns)
    tv1["show"] = "headings"
    for column in tv1["columns"]:
        tv1.heading(column, text=column) # let the column heading = column name

    df_rows = df.to_numpy().tolist() # turns the dataframe into a list of lists
    for row in df_rows:
        tv1.insert("", "end", values=row) # inserts each list into the treeview. For parameters see https://docs.python.org/3/library/tkinter.ttk.html#tkinter.ttk.Treeview.insert
    return None


def clear_data():
    tv1.delete(*tv1.get_children())
    return None



# ---------------------------------- OpenpyXl to loop through Band names ----------------------------------

try:

    wb=Workbook()

    wb_path="D:/Import Files/Band.xlsx"
    wbk = openpyxl.load_workbook(wb_path)
    ws = wbk['Band']
    Range_excel=ws['A2':'A1000']
    bnd_dt=[]

    for cell in Range_excel: 
        for x in cell:
            if x.value != None:
                bnd_dt.append(x.value)
            else:
                pass


    wb=Workbook()

    wb_path="D:/Import Files/Band.xlsx"
    wbk = openpyxl.load_workbook(wb_path)
    ws = wbk['Band']

except:
    pass




driver = None
global relative_path

# ------------- for Chromedriver, adjustment to add it in the exe --------------------

def resource_path1(relative_path):
    try:
        base_path = sys.MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return base_path + relative_path

# -------------- for Image, adjustment to add it in the exe -------------------------

def resource_path_logo(relative_path_logo):
    try:
        base_path_logo = sys._MEIPASS
        print(base_path_logo)
    except Exception:
        base_path_logo = os.path.dirname(__file__)
        print(base_path_logo)
    return base_path_logo + relative_path_logo


 #pyi-makespec Py_Selenium.py --onefile --noconsole --add-binary "driver\chromedriver.exe;driver\" --add-data "Img\Dbox4.png;Img\"  --name Core_DCT
 #pyinstaller --clean Core_DCT.spec  


def on_open():
    global driver

    if not driver:
        print(os.path.dirname(__file__))
        driver = selenium.webdriver.Chrome(resource_path1("/driver/chromedriver.exe"))
        url = WebLink.get()
        driver.get(url)
        driver.maximize_window()


def on_close():
    global driver

    if driver:
        driver.close()
        driver = None



def UserLogin ():

    global driver

    # Admin's login
    driver.find_element_by_id("UserLogin_username").send_keys(username1.get())
    driver.find_element_by_id("UserLogin_password").send_keys(password1.get())
    driver.find_element_by_id("login-submit").click()
    driver.implicitly_wait(10)
    

    # How are you feeling today?

    try:
        driver.find_element_by_xpath('//*[@id="pulse_form"]/div/div/div')
        driver.find_element_by_xpath('//*[@id="5"]').click()
        driver.find_element_by_xpath('//*[@id="plus-status-btn"]').click()
    except:
        pass

    #Click on the user's profile pic and switch to admin
    driver.find_element_by_xpath('//*[@id="dasboard-bigheader"]/header/div[4]/ul/li[3]/div/div/img').click()
    driver.find_element_by_xpath('//*[@id="dasboard-bigheader"]/header/div[4]/ul/li[3]/div/ul/li[2]/a').click()
    driver.implicitly_wait(30)

    time.sleep(5)

    #https://training2.darwinbox.in/settings/accountdetails
    driver.get(WebLink.get() + '/settings/accountdetails')

    time.sleep(5)

    # ------------------  Asking user if GC needs to be created. Answer == true, GC related input boxes will be visible else not visible -------------------

    if UserGCQuestion == True:
        
        try:

            driver.get(WebLink.get() + '/settings/company')

            #To clear the default valeus present in the form
            driver.find_element_by_xpath('//*[@id="Tenants_tenant_name"]').clear()
            driver.find_element_by_xpath('//*[@id="Tenants_tenant_name"]').send_keys(GC_name.get())

            # driver.find_element_by_xpath('//*[@id="TenantProfile_tenant_code"]').clear()
            # driver.find_element_by_xpath('//*[@id="TenantProfile_tenant_code"]').send_keys(Tenant_Code)

            driver.find_element_by_xpath('//*[@id="TenantProfile_tenant_shortname"]').clear()
            driver.find_element_by_xpath('//*[@id="TenantProfile_tenant_shortname"]').send_keys(GC_shortName1.get())

            #time.sleep(5)

            # TO find dropdown and then select the Country
            #driver.find_element_by_xpath('//*[@id="manage-tenant-form"]/div[6]/div[5]/div').click()
            Sel = Select(driver.find_element_by_xpath('//*[@id="country_add"]'))
            Sel.select_by_visible_text(GC_country1.get())

            #State
            driver.find_element_by_xpath('//*[@id="manage-tenant-form"]/div[6]/div[6]/div/input').send_keys(GC_State1.get())
            driver.implicitly_wait(20)

            #City
            driver.find_element_by_xpath('//*[@id="manage-tenant-form"]/div[7]/div[1]/div/input').send_keys(GC_city1.get())
            driver.implicitly_wait(20)

            # Save the information
            driver.find_element_by_xpath('//*[@id="company_update_btn"]').click()
            driver.implicitly_wait(30)
            
        except:
            pass
    else:
        pass

    # ------------------------- Storing aliases and clearing them ------------------------------

    # Department alias
    try:
        driver.get(WebLink.get() + '/settings/company/companyalias')
        global StoringDeptAlias 
        StoringDeptAlias = driver.find_element_by_xpath('//*[@id="AliasSettings_department"]').text
        time.sleep(4)
        driver.find_element_by_xpath('//*[@id="AliasSettings_department"]').clear()
    except:
        pass

    # Desgination alias
    try:

        global StoringDesgAlias
        StoringDesgAlias = driver.find_element_by_xpath('//*[@id="AliasSettings_designations"]').text
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_designations"]').clear()
    except:
        pass

    # Desgination title
    try:
        global StoringDesgTitleAlias
        StoringDesgTitleAlias = driver.find_element_by_xpath('//*[@id="AliasSettings_designation_title"]').text
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_designation_title"]').clear()
    except:
        pass

    # Grade alias 
    try:

        global StoringGradeAlias
        StoringGradeAlias = driver.find_element_by_xpath('//*[@id="AliasSettings_grade"]').text
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_grade"]').clear()
    except:
        pass

    # Band Alias 
    try:

        global StoringBandAlias
        StoringBandAlias = driver.find_element_by_xpath('//*[@id="AliasSettings_band"]').text
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_band"]').clear()

    except:
        pass
    
    
    #Save 
    try:
        driver.find_element_by_xpath('//*[@id="leave_settings_create_btn"]').click()
    except:
        pass   
# --------------------- Access the Company Import Page and Upload----------------------------------

def Upload_file_1():

    #Defining wait variable which can be used in wait.until    
    wait = WebDriverWait(driver, 30)

    # Cost center

    try:

        time.sleep(4)
        driver.get(WebLink.get() + '/Importnewclient/costCenters')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Cost_center)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Location Type

    try:

        driver.get(WebLink.get() + '/Importnewclient/locationType')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Loc_type)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # City Type

    try:

        driver.get(WebLink.get() + '/Importnewclient/cityType')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(City_type)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Grade

    try:
        driver.get(WebLink.get() + '/Importnewclient/grade')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Grade_upload)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Contribution level
    # If user selects to add contribution level then this will run
    # is_selected will check if the checkbox is already 'checked' else will go ahead and select the option

    if CheckBox_Contribution_var.get() == 1:

        try:
            driver.get(WebLink.get() + '/settings/employees/tenprofzz')
            driver.execute_script("window.scrollTo(0, 1600)") 
            time.sleep(4)

            if driver.find_element_by_id("TenantProfile_neev_level_allowed").is_selected() == False:
                time.sleep(4)
                driver.find_element_by_id("TenantProfile_neev_level_allowed").click()
                driver.execute_script("window.scrollTo(0, -500)")
                time.sleep(4)

                driver.find_element_by_xpath("/html/body/div[2]/div/section/div/div/div/form/div[2]/div/input").click()
                time.sleep(4)
            
            else:
                pass

            driver.get(WebLink.get() + '/Importnewclient/contributionLevel')
            driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Contrib_level)
            driver.implicitly_wait(20)
            element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
            time.sleep(4)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
            #driver.find_element_by_name('upload').click()
            driver.implicitly_wait(20)

        except:
            pass

    else:
        pass

    # Business unit

    try:

        driver.get(WebLink.get() + '/Importnewclient/businessUnit')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Business_unit_up)
        # ------------------ Using wait until and EC component ---------------------------
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Division
    try:

        time.sleep(4)
        #driver.switch_to_window(driver.window_handles[1])
        driver.get(WebLink.get() + '/Importnewclient/division')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Division_up)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Department master
    try:

        time.sleep(4)
        driver.get(WebLink.get() + '/Importnewclient/department')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Dept_master)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Department mapping
    try:

        time.sleep(4)
        driver.get(WebLink.get() + '/Importnewclient/departmentmapping')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Dept_mapping)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Locations

    try:

        driver.get(WebLink.get() + '/Importnewclient/Location')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Locations_upload)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Band

    try:

        for i in bnd_dt:
            wait = WebDriverWait(driver, 30)
            driver.get('https://training2.darwinbox.in/settings/company/bands')
            time.sleep(4)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a[1]').click()
            driver.find_element_by_xpath('//*[@id="UserBand_band_name"]').send_keys(i)
            driver.implicitly_wait(5)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[2]/div/input').click()
            time.sleep(5)
    
    except:
        pass

    # Job level

    if CheckBox_Joblevel_var == 1:

        try:

            driver.get(WebLink.get() + '/Importnewclient/jobLevel')
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(JobLevel_upload)
            driver.implicitly_wait(20)
            element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
            time.sleep(4)
            driver.find_element_by_xpath('//*[@id="upload_import_file"]/div/input').click()
            #driver.find_element_by_name('upload').click()
            driver.implicitly_wait(20)

        except:
            pass

    else:

        try:

            driver.get(WebLink.get() + '/import/gradeimport')
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="upload_file[]"]').send_keys(Grade_in_designation)
            driver.implicitly_wait(20)

            driver.find_element_by_xpath('//*[@id="has_header_fields"]').click()

            element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
            time.sleep(4)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/div[1]/div/div/div[6]/form/div/input').click()
            #driver.find_element_by_name('upload').click()
            driver.implicitly_wait(20)

            
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/div/div/div/div[9]/div[2]/div[5]/div/table/thead/tr/th[1]/select').send_keys('Designation Code')
            driver.implicitly_wait(20)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/div/div/div/div[9]/div[2]/div[5]/div/table/thead/tr/th[2]/select').send_keys('Grade')
            driver.implicitly_wait(20)

            time.sleep(3)

            driver.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/div/div/div/div[6]/form/div/input').click()
            time.sleep(3)

            try:
                driver.execute_script("window.scrollTo(0, 300)") 
                driver.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/div[4]/input[1]').click()
                time.sleep(3)
            except:
                pass


        except:
            pass


    # Functional area
    
    try:

        driver.get(WebLink.get() + '/Importnewclient/functionalArea')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Func_area)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    # Functional area mapping

    try:

        driver.get(WebLink.get() + '/Importnewclient/famapping')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Func_area_mapping)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('/html/body/div[2]/div/section/div[1]/div[1]/div/div/form/input[2]').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass

    #Designation

    try:

        driver.get(WebLink.get() + '/import/asyncImports/type/designation')
        driver.find_element_by_xpath('//*[@id="upload_file[]"]').send_keys(Designation)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('//*[@id="upload_import_file"]/div/input').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass
    
    # Designation Name

    try:

        driver.get(WebLink.get() + '/import/asyncImports/type/designationname')
        driver.find_element_by_xpath('//*[@id="upload_file[]"]').send_keys(Designation_name)
        driver.implicitly_wait(20)
        element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload'))))
        time.sleep(4)
        driver.find_element_by_xpath('//*[@id="upload_import_file"]/div/input').click()
        #driver.find_element_by_name('upload').click()
        driver.implicitly_wait(20)

    except:
        pass


    # Designation Location

    try:

        time.sleep(4)
        driver.get(WebLink.get() + '/import/designationLocation')
        driver.find_element_by_xpath('//*[@id="csvdata"]').send_keys(Designation_loc)
        time.sleep(4)
        driver.implicitly_wait(20)
    #   element = wait.until(EC.element_to_be_clickable(((By.NAME,'upload')))) --- not required as this page doesnt have a submit or next button present

    except:
        pass

    #except:
        #pass

    #------------------------------- code to select from Deignation's drop down ------------------------------------

    try:

        driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/table/tbody/tr[1]/td[1]/select').send_keys(Dept)
        driver.implicitly_wait(3)
        driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/table/tbody/tr[2]/td[1]/select').send_keys(Desg)
        driver.implicitly_wait(3)
        driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/table/tbody/tr[3]/td[1]/select').send_keys(Desg_code)
        driver.implicitly_wait(3)
        driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/table/tbody/tr[4]/td[1]/select').send_keys(Numb_of_position)
        driver.implicitly_wait(3)
        driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/table/tbody/tr[8]/td[1]/select').send_keys(Not_period)
        driver.implicitly_wait(3)
        driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/table/tbody/tr[10]/td[1]/select').send_keys(Funct_area)
        time.sleep(2)

        driver.execute_script("window.scrollTo(0, -500)") 

        time.sleep(5)

        driver.find_element_by_xpath('//*[@id="col_map"]/div/input').click()

        time.sleep(5)

    except:
        pass

    # ----------------------- Aliases which was deleted earlier will be added again and saved ------------------------

    # Department alias adding again
    try:
        driver.get(WebLink.get() + '/settings/company/companyalias')
        time.sleep(4)
        driver.find_element_by_xpath('//*[@id="AliasSettings_department"]').send_keys(StoringDeptAlias)
    except:
        pass

    # Desgination alias adding again
    try:
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_designations"]').send_keys(StoringDesgAlias)
    except:
        pass

    # Desgination title adding again
    try:
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_designation_title"]').send_keys(StoringDesgTitleAlias)
    except:
        pass

    # Grade alias adding again
    try:

        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_grade"]').send_keys(StoringGradeAlias)
    except:
        pass

    # Band Alias adding again
    try:
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="AliasSettings_band"]').send_keys(StoringBandAlias)
    except:
        pass

    try:
        driver.find_element_by_xpath('//*[@id="leave_settings_create_btn"]').click()
    except:
        pass

def Help_window():
    messagebox.showinfo(title="How to use this executable",message=InfoForTHeUser)
    


# --------------------------------------------------------------------------------------------------------------
# ---------------------------------------------- Tkinter -------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------

InfoForTHeUser ="Please make sure that you have used the Automated DCT to take the export and the exported CSVs are present in 'D: Import files' folder. \r\n\r\nIf Client's GC information is added then kindly select 'No' or select 'Yes' if you want to change/update the GC info \r\n \r\n \r\n Some points to remeber -> \r\n\r\n 1. Please make sure that you are adding the entire instance link. that is make sure to include 'https://' also \r\n\r\n 2. Please add the correct user ID and Password \r\n\r\n 3. 'Delays' are added purposely so as to let the website load all its elements. \r\n\r\n 4. Please select Contribution and Job level checkboxes if you want to upload them. Note : If user selects not to upload Job level then Grade in designation csv will be uploaded. \r\n \r\n\r\n Information about buttons -> \r\n 1. Selenium Open -> Will open a new Google chrome window and maximize it automatically \r\n\r\n 2. Login & Admin -> User ID and Password added will be used to login and switch from employee's profile to Admin's \r\n\r\n  3. Upload Files  -> CSV exports from D -> Import files folder will be uploaded \r\n\r\n 4. Selenium close -> will close this program\r\n"

driver = None

root  = tk.Tk()
#new_root = tk.Tk()

#Logo = resource_path("Logo.png")
# Logo url -> share point public (vis for eve in org)

img = PhotoImage(file=resource_path_logo('/Img/Dbox4.png'))

Label(root,image=img,height=60).grid(sticky=NE,padx=20,pady=20)

root.title('DarwinBox')
#width then hight
root.geometry('950x650+150+150')

root['bg'] = '#5252ff'


UserGCQuestion = messagebox.askyesno(title="Group company",message="Do you want to Add / update GC information?")


if UserGCQuestion == True :

    tk.Label(root,text="Client Instance / Website link ->",width=25).grid(row=3,column=1,padx=10,pady=5)
    WebLink = StringVar()
    name1 = tk.Entry(root, textvariable=WebLink,width=30)
    name1.grid(row=3,column=2,padx=5,pady=5)

    tk.Label(root,text="User ID / Email ID ->",activebackground='white',width=25).grid(row=3,column=3,padx=10,pady=3)
    username1 = StringVar()
    name2 = tk.Entry(root, textvariable=username1,width=30)
    name2.grid(row=3,column=4,padx=5,pady=3)

    tk.Label(root,text="Password ->",width=25).grid(row=4,column=1,padx=10,pady=3)
    password1 = StringVar()
    name3 = tk.Entry(root, textvariable=password1,show="*",width=30)
    name3.grid(row=4,column=2,padx=5,pady=3)

    tk.Label(root,text="GC Name ->",width=25).grid(row=4,column=3,padx=10,pady=3)
    GC_name = StringVar()
    name4 = tk.Entry(root, textvariable=GC_name,width=30)
    name4.grid(row=4,column=4,padx=5,pady=3)

    tk.Label(root,text="GC Shortname ->",width=25).grid(row=5,column=1,padx=10,pady=3)
    GC_shortName1 = StringVar()
    name6 = tk.Entry(root, textvariable=GC_shortName1,width=30)
    name6.grid(row=5,column=2,padx=5,pady=3)

    tk.Label(root,text="GC Country ->",width=25).grid(row=5,column=3,padx=10,pady=3)
    GC_country1 = StringVar()
    name6 = tk.Entry(root, textvariable=GC_country1,width=30)
    name6.grid(row=5,column=4,padx=5,pady=3)

    tk.Label(root,text="GC State ->",width=25).grid(row=6,column=1,padx=10,pady=3)
    GC_State1 = StringVar()
    name7 = tk.Entry(root, textvariable=GC_State1,width=30)
    name7.grid(row=6,column=2,padx=5,pady=3)

    tk.Label(root,text="GC City ->",width=25).grid(row=6,column=3,padx=10,pady=3)
    GC_city1 = StringVar()
    name8 = tk.Entry(root, textvariable=GC_city1,width=30)
    name8.grid(row=6,column=4,padx=5,pady=3)

    
else :
 
    tk.Label(root,text="Client Instance / Website link ->",width=20).grid(row=3,column=1,padx=10,pady=10)
    WebLink = StringVar()
    name1 = tk.Entry(root, textvariable=WebLink,width=100)
    name1.grid(row=3,column=2,padx=5,pady=10,columnspan=3)

    tk.Label(root,text="User ID / Email ID ->",activebackground='white',width=20).grid(row=4,column=1,padx=10,pady=5)
    username1 = StringVar()
    name2 = tk.Entry(root, textvariable=username1,width=35)
    name2.grid(row=4,column=2,padx=5,pady=10)

    tk.Label(root,text="Password ->",width=20).grid(row=4,column=3,padx=10,pady=5)
    password1 = StringVar()
    name3 = tk.Entry(root, textvariable=password1,show="*",width=35)
    name3.grid(row=4,column=4,padx=5,pady=10)
    

CheckBox_Contribution_var = IntVar()
CheckBox_Contribution1 = tk.Checkbutton(root, text="Is Contribution level applicable ?", variable=CheckBox_Contribution_var, onvalue=1, offvalue=0,activebackground='blue').grid(row=7,column=1,padx=10,pady=5,columnspan=2)

CheckBox_Joblevel_var= IntVar()
CheckBox_Job3 = tk.Checkbutton(root, text="Is job level applicable ?", variable=CheckBox_Joblevel_var, onvalue=1, offvalue=0,activebackground='blue').grid(row=7,column=3,padx=10,pady=5,columnspan=2)

b1 = tk.Label(root,text="*** Please provide all the inputs ***",width=110,background='#5252ff').grid(row=8,column=1,padx=10,pady=10,columnspan=4)

b2 = tk.Button(root, text='Select file', command=select_file,width=40,relief=RAISED,activebackground='Grey',background='#285DC0',fg='white').grid(row=9,column=1,padx=10,pady=5,columnspan=4)

b10 = tk.Button(root, text='Check Errors', command=run_excel_macro_validation,width=40,relief=RAISED,activebackground='Grey',bg='#285DC0',fg='white').grid(row=10,column=1,padx=10,pady=5,columnspan=2)
b11 = tk.Button(root, text='Show validation errors', command=EntireTree,width=40,relief=RAISED,activebackground='Grey',bg='#285DC0',fg='white').grid(row=10,column=3,padx=10,pady=5,columnspan=2)
b12 = tk.Button(root, text='Resolve all errors', command=run_validation_Allresolved_macro,width=40,relief=RAISED,activebackground='Grey',bg='#285DC0',fg='white').grid(row=11,column=1,padx=10,pady=10,columnspan=4)

#Validate data button -> validation ->Error report

b3 = tk.Button(root, text='Generate Core Master Imports', command=run_excel_macro,width=40,relief=RAISED,activebackground='Grey',background='black',fg='white').grid(row=12,column=1,padx=10,pady=5,columnspan=4)

b4 = tk.Label(root,text="",background='#5252ff').grid(row=13,column=1,padx=10,pady=1,columnspan=4)
 
b5 = tk.Button(root, text='Chrome Open', command=on_open,width=40,relief=RAISED,activebackground='Grey').grid(row=14,column=1,padx=10,pady=5,columnspan=4)

b6 = tk.Button(root, text='Login & Admin', command=UserLogin,width=40,relief=RAISED,activebackground='Grey').grid(row=15,column=1,padx=10,pady=5,columnspan=4)

#b3 = tk.Button(root, text='Account details', command=GC_Details,width=40).grid(row=7,column=1,padx=10,pady=5,columnspan=3)

b7 = tk.Button(root, text='Upload Files', command=Upload_file_1,width=40,relief=RAISED,activebackground='Grey').grid(row=16,column=1,padx=10,pady=5,columnspan=4)

#b5 = tk.Button(root, text='Upload File_2', command=upload_file_2,width=40).grid(row=9,column=1,padx=10,pady=5,columnspan=3)

b8 = tk.Button(root, text='Chrome Close', command=on_close,width=40,relief=RAISED,activebackground='Grey').grid(row=17,column=1,padx=10,pady=5,columnspan=4)

b9 = tk.Button(root, text='Help', command=Help_window,width=40,relief=RAISED,activebackground='Grey').grid(row=18,column=1,padx=10,pady=15,columnspan=4)




root.mainloop()



