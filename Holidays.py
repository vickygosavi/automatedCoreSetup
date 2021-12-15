
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from openpyxl.descriptors.base import DateTime, String
from openpyxl.worksheet.pagebreak import Break
import selenium.webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common import keys
from selenium.webdriver.support import select
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import *
import os
import sys
import xlwings as xw
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from xlwings.main import Sheet, Sheets


driver = None
global relative_path

# ------------- for Chromedriver, adjustment to add it in the exe --------------------

def resource_path1(relative_path):
    try:
        base_path = sys.MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return base_path + relative_path

# -------------- for Image, adjustment to add it in the exe -------------------------

def resource_path_logo(relative_path_logo):
    try:
        base_path_logo = sys._MEIPASS
        print(base_path_logo)
    except Exception:
        base_path_logo = os.path.dirname(__file__)
        print(base_path_logo)
    return base_path_logo + relative_path_logo


 #pyi-makespec Holidays.py --onefile --noconsole --add-binary "driver\chromedriver.exe;driver\" --add-data "Img\Dbox4.png;Img\"  --name Holidays
 #pyinstaller --clean Holidays.spec

def on_open():
    global driver

    if not driver:
        print(os.path.dirname(__file__))
        driver = selenium.webdriver.Chrome(resource_path1("/driver/chromedriver.exe"))
        url = WebLink.get()
        driver.get(url)
        driver.maximize_window()


def on_close():
    global driver

    if driver:
        driver.close()
        driver = None


def UserLogin ():

    global driver

    # Admin's login
    driver.find_element_by_id("UserLogin_username").send_keys(username1.get())
    driver.find_element_by_id("UserLogin_password").send_keys(password1.get())
    driver.find_element_by_id("login-submit").click()
    driver.implicitly_wait(3)

    # How are you feeling today?
    try:
        driver.find_element_by_xpath('//*[@id="pulse_form"]/div/div/div')
        driver.find_element_by_xpath('//*[@id="5"]').click()
        driver.find_element_by_xpath('//*[@id="plus-status-btn"]').click()
        time.sleep(5)
    except:
        pass

    #Click on the user's profile pic and switch to admin
    driver.find_element_by_xpath('//*[@id="dasboard-bigheader"]/header/div[4]/ul/li[3]/div/div/img').click()
    driver.find_element_by_xpath('//*[@id="dasboard-bigheader"]/header/div[4]/ul/li[3]/div/ul/li[2]/a').click()
    driver.implicitly_wait(3)
    time.sleep(2)

#------------------------------Leaves Process Mapping Automation---------------------------------------------------

def Holidays ():

   global driver
   wb=Workbook()
   wb_path= r"D:\Leaves-DCT_v3.xlsx"
   wbk = openpyxl.load_workbook(wb_path)

   a = ["6. Holidays"]

   for i in a :
      ws = wbk[i]
      if ws['C7'].value != None : 

         driver.get(WebLink.get() + '/settings/leaves/holidays')

         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[1]/input').clear()
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[1]/input').send_keys(ws['C7'].value)
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[2]/input').click()
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[2]/input').send_keys(ws['D7'].value)
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[2]/input').send_keys(Keys.ENTER)
         if ws['E7'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[3]/div/input').click()
         if ws['F7'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[4]/select').send_keys(ws['F7'].value)
         if ws['G7'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[5]/div/ul/li/input').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[5]/div/ul/li/input').send_keys(ws['G7'].value)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr/td[5]/div/ul/li/input').send_keys(Keys.ENTER)

         HCount = ["2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34"]
         CellNo = ["8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40"]

         for x,y in enumerate(HCount) :
            ActualCellNo=CellNo[x]
            
            if ws["C" + ActualCellNo].value != None : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/a').click()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[1]/input').send_keys(ws["C" + ActualCellNo].value)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[2]/input').clear()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[2]/input').send_keys(ws['D'+ ActualCellNo].value)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[2]/input').send_keys(Keys.ENTER)
               if ws['E'+ ActualCellNo].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[3]/div/input').click()
               if ws['F'+ ActualCellNo].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[4]/select').send_keys(ws['F'+ ActualCellNo].value)
               if ws['G'+ ActualCellNo].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[5]/div/ul/li/input').click()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[5]/div/ul/li/input').send_keys(ws['G'+ ActualCellNo].value)
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/form/div/div[3]/table/tbody/tr['+y+']/td[5]/div/ul/li/input').send_keys(Keys.ENTER)
      driver.execute_script("window.scrollTo(0, 0)")
      time.sleep(2)
      driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/div[1]/div[2]/div/button').click()
time.sleep(2)

# --------------------------------------------------------------------------------------------------------------
# ---------------------------------------------- Tkinter -------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------

InfoForTHeUser ="Please make sure that you have used 'Leaves Process Mapping_v3' to take the data and it is present in 'D:' Drive."

driver = None

root  = tk.Tk()

img = PhotoImage(file=resource_path_logo('/Img/Dbox4.png'))

Label(root,image=img,height=60).grid(sticky=NE,padx=20,pady=20)

root.title('Holidays')
#width then hight
root.geometry('600x425+50+50')

root['bg'] = '#5252ff'

tk.Label(root,text="Client Instance / Website link ->").grid(row=3,column=1,padx=5,pady=5)
WebLink = StringVar()
name1 = tk.Entry(root, textvariable=WebLink,width=30)
name1.grid(row=3,column=2,padx=5,pady=5)

tk.Label(root,text="User ID / Email ID ->",activebackground='white').grid(row=4,column=1,padx=5,pady=5)
username1 = StringVar()
name2 = tk.Entry(root, textvariable=username1,width=30)
name2.grid(row=4,column=2,padx=5,pady=5)

tk.Label(root,text="Password ->").grid(row=5,column=1,padx=5,pady=5)
password1 = StringVar()
name3 = tk.Entry(root, textvariable=password1,show="*",width=30)
name3.grid(row=5,column=2,padx=5,pady=5)

b7 = tk.Label(root,text="*** Please provide all the inputs ***",width=55).grid(row=8,column=1,padx=5,pady=5,columnspan=4)

b6 = tk.Label(root,text="",background='#5252ff').grid(row=11,column=1,padx=10,pady=1,columnspan=4)
 
b1 = tk.Button(root, text='Chrome Open', command=on_open,width=40,relief=RAISED,activebackground='Grey').grid(row=12,column=1,padx=5,pady=5,columnspan=4)

b2 = tk.Button(root, text='Login & Admin', command=UserLogin,width=40,relief=RAISED,activebackground='Grey').grid(row=13,column=1,padx=5,pady=5,columnspan=4)

b3 = tk.Button(root, text='Create Holidays', command=Holidays,width=40,relief=RAISED,activebackground='Grey').grid(row=14,column=1,padx=5,pady=5,columnspan=4)

b5 = tk.Button(root, text='Chrome Close', command=on_close,width=40,relief=RAISED,activebackground='Grey').grid(row=15,column=1,padx=10,pady=5,columnspan=4)

root.mainloop()
