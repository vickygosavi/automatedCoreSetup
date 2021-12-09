
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from openpyxl.descriptors.base import DateTime, String
from openpyxl.worksheet.pagebreak import Break
#import tkinter
#from typing_extensions import ParamSpecKwargs
import selenium.webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common import keys
from selenium.webdriver.support import select
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import *
import os
import sys
import xlwings as xw
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from xlwings.main import Sheet, Sheets



driver = None
global relative_path

# ------------- for Chromedriver, adjustment to add it in the exe --------------------

def resource_path1(relative_path):
    try:
        base_path = sys.MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return base_path + relative_path

# -------------- for Image, adjustment to add it in the exe -------------------------

def resource_path_logo(relative_path_logo):
    try:
        base_path_logo = sys._MEIPASS
        print(base_path_logo)
    except Exception:
        base_path_logo = os.path.dirname(__file__)
        print(base_path_logo)
    return base_path_logo + relative_path_logo


 #pyi-makespec Add_Leave_Policy.py --onefile --noconsole --add-binary "driver\chromedriver.exe;driver\" --add-data "Img\Dbox4.png;Img\"  --name Add_Leave_Policy
 #pyinstaller --clean Add_Leave_Policy.spec

def on_open():
    global driver

    if not driver:
        print(os.path.dirname(__file__))
        driver = selenium.webdriver.Chrome(resource_path1("/driver/chromedriver.exe"))
        url = WebLink.get()
        driver.get(url)
        driver.maximize_window()


def on_close():
    global driver

    if driver:
        driver.close()
        driver = None


def UserLogin ():

    global driver

    # Admin's login
    driver.find_element_by_id("UserLogin_username").send_keys(username1.get())
    driver.find_element_by_id("UserLogin_password").send_keys(password1.get())
    driver.find_element_by_id("login-submit").click()
    driver.implicitly_wait(3)

    # How are you feeling today?
    try:
        driver.find_element_by_xpath('//*[@id="pulse_form"]/div/div/div')
        driver.find_element_by_xpath('//*[@id="5"]').click()
        driver.find_element_by_xpath('//*[@id="plus-status-btn"]').click()
        time.sleep(5)
    except:
        pass

    #Click on the user's profile pic and switch to admin
    driver.find_element_by_xpath('//*[@id="dasboard-bigheader"]/header/div[4]/ul/li[3]/div/div/img').click()
    driver.find_element_by_xpath('//*[@id="dasboard-bigheader"]/header/div[4]/ul/li[3]/div/ul/li[2]/a').click()
    driver.implicitly_wait(3)
    time.sleep(2)

#------------------------------Leaves Process Mapping Automation---------------------------------------------------

def Add_Leave_Policy ():

   global driver
   wb=Workbook()
   wb_path= r"D:\Leaves Process Mapping_v3.xlsx"
   wbk = openpyxl.load_workbook(wb_path)

   a = ["3.1 PL Policy","3.2 EL Policy","3.3 CL Policy","3.4 SL Policy","3.5 AL Policy","3.6 MtL Policy","3.7 PtL Policy","3.8 SbL Policy","3.9 ML Policy","3.10 GL Policy","3.11 Custom Leave 1","3.12 Custom Leave 2","3.13 Custom Leave 3"]
 
   for i in a :
      ws = wbk[i]
      if ws['G7'].value != None : 

         driver.get(WebLink.get() + '/settings/leaves')
                                      
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a').click()
         AssignmentType=ws['G14']
         if AssignmentType.value == "Company Wise": 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[1]/div/div[1]/input').click()
         if ws['G15'].value != None : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[3]/div/select').send_keys(ws['G15'].value)
            time.sleep(3)

         if AssignmentType.value == "Assignment Framework" : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[1]/div/div[2]/input').click()
            if ws['G16'].value != None :
               AssTypeClick = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[2]/div/div/ul/li/input').click()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[2]/div/div/ul/li/input').send_keys(ws['G16'].value)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[2]/div/div/ul/li/input').send_keys(Keys.ENTER)

         driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[4]/input').send_keys(ws['G7'].value)

         if ws['G9'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[7]/textarea').send_keys(ws['G9'].value)

         if ws['G42'].value != None : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[8]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[8]/input').send_keys(ws['G42'].value)
            time.sleep(2)

         if AssignmentType.value == "Company Wise":  
        
            if ws['G18'].value == "AND" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[10]/div[1]/div[1]/input').click()
            if ws['G18'].value == "OR"  : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[10]/div[1]/div[2]/input').click()
            time.sleep(1)
#'Restriction (Department, Employee Type Or Location)'
            if ws['G17'].value != None : 
      
               Search = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[10]/div[2]/div/ul/li/input').click()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[10]/div[2]/div/ul/li/input').send_keys(ws['G17'].value)
               time.sleep(1)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[10]/div[2]/div/ul/li/input').send_keys(Keys.ENTER)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[10]/div[2]/div/ul/li/input').send_keys(Keys.TAB)
               
         if ws['G99'].value != None : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[11]/input').clear()
            
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[11]/input').send_keys(ws['G99'].value)
            time.sleep(1)
         if ws['G25'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[12]/select').send_keys(ws['G25'].value)
            time.sleep(1)

            if ws['G25'].value == "Continuous Cycle" :
               time.sleep(1)
               if ws['G26'].value == "YES" : 
                  time.sleep(1)
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[15]/div[1]/div/input').click()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[15]/div[2]/input').send_keys(ws['G27'].value)

            if ws['G25'].value =="Custom Cycle" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[14]/select').send_keys(ws['G28'].value)
         if ws['G164'].value != None : 
            if ws['G164'].value =="YES" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[5]/div[16]/div/input').click()
         if ws['G97'].value != None :
            if ws['G97'].value =="NO" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.TAB)
            if ws['G97'].value =="Monday" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            if ws['G97'].value =="Tuesday" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            if ws['G97'].value =="Wednesday" :  
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            if ws['G97'].value =="Thursday" :  
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            if ws['G97'].value =="Friday" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            if ws['G97'].value =="Saturday" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            if ws['G97'].value =="Sunday" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.DOWN)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            time.sleep(1)

         if ws['G162'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[2]/input').send_keys(ws['G162'].value)
         if ws['G107'].value != None :
            if ws['G107'].value =="YES" :
               time.sleep(1)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[3]/div/input').click() 
               if ws['G108'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[4]/div[1]/input').send_keys(ws['G108'].value)
               if ws['G109'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[4]/div[2]/input').send_keys(ws['G109'].value)
               if ws['G110'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[4]/div[3]/input').send_keys(ws['G110'].value)

         if ws['G148'].value != None :
            if ws['G148'].value =="YES" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[5]/div/input').click()
         if ws['G30'].value != None :
            if ws['G30'].value == "For All" or "Female Only" or "Male Only" or "Do not want to disclose Only" or "Transgender Only" or "Other Only" :
               Gender = ws['G30'].value
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[6]/div/a/span').click()
               time.sleep(1)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[6]/div/div/div/input').send_keys(Gender)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[6]/div/div/div/input').send_keys(Keys.ENTER)

         if ws['G92'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[16]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[16]/input').send_keys(ws['G92'].value)
         
         if ws['G134'].value != None:
            if ws['G134'].value == "Custom Months" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[18]/div[1]/input').click()
               if ws['G135'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[19]/input').clear()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[19]/input').send_keys(ws['G135'].value)

            if ws['G134'].value == "Employee Probation Period" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[6]/div[18]/div[2]/input').click()
            if ws['G134'].value == "NA" : pass
         if ws['G101'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[1]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[1]/input').send_keys(ws['G101'].value)
            
         if ws['G141'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[2]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[2]/input').send_keys(ws['G141'].value)
            
         if ws['G142'].value == "YES" :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[3]/div/input').click()
            
         if ws['G139'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[4]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[4]/input').send_keys(ws['G139'].value)
         if ws['G88'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[5]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[5]/input').send_keys(ws['G88'].value)
         if ws['G103'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[7]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[7]/input').send_keys(ws['G103'].value)
         if ws['G105'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[8]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[8]/input').send_keys(ws['G105'].value)
         if ws['G150'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[9]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[9]/input').send_keys(ws['G150'].value)
         if ws['G32'].value == "YES" :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[10]/div/input').click()
         if ws['G32'].value == "NO" : 
            if ws['G34'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[11]/div/input').click()
            if ws['G36'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[12]/div/input').click()
         if ws['G144'].value == "YES" :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[13]/div/input').click()
         if ws['G137'].value == "YES" :   
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[14]/div/input').click()
         if ws['G38'].value == "YES" :  
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[16]/div/input').click()
         if ws['G154'].value != None :  
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[17]/select').send_keys(ws['G154'].value)
         if ws['G156'].value != None :   
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[19]/div[1]/select').send_keys(ws['G156'].value)
            if ws['G158'].value == "YES" :  
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[19]/div[2]/input').click()
               time.sleep(1)
            if ws['G159'].value != None :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[19]/div[3]/div[1]/input').clear()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[19]/div[3]/div[1]/input').send_keys(ws['G159'].value)
            if ws['G160'].value != None :   
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[19]/div[3]/div[2]/select').send_keys(ws['G160'].value)
         if ws['G156'].value == None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[7]/div[18]/div/input[2]').click()
         if ws['G44'].value == "YES" : 
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[1]/input').click()
            if ws['G45'].value == "AND" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[1]/div[1]/input').click()
               if ws['G46'].value != None : 
                  Search1=driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[2]/div/ul/li/input').click()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[2]/div/ul/li/input').send_keys(ws['G46'].value)
                  time.sleep(1)
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[2]/div/ul/li/input').send_keys(Keys.ENTER)
               if ws['G47'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[3]/input').send_keys(ws['G47'].value)

            if ws['G45'].value == "OR" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[1]/div[2]/input').click()
               if ws['G46'].value != None : 
                  Search2=driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[2]/div/ul/li/input').click()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[2]/div/ul/li/input').send_keys(ws['G46'].value)
                  time.sleep(1)
                  Search2=driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[2]/div/ul/li/input').send_keys(Keys.ENTER)
               if ws['G47'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[1]/td[3]/input').send_keys(ws['G47'].value)
         
            if ws['G49'].value == "AND" or "OR" : 
               time.sleep(1)
               if ws['G49'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/div/a').click()
               if ws['G49'].value == "AND" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[1]/div[1]/input').click()
                  if ws['G50'].value != None : 
                     Search3 = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[2]/div/ul/li/input').click()
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[2]/div/ul/li/input').send_keys(ws['G50'].value)
                     time.sleep(1)
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[2]/div/ul/li/input').send_keys(Keys.ENTER)
                  if ws['G51'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[3]/input').send_keys(ws['G51'].value)
            
                  if ws['G49'].value == "OR" :
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[1]/div[2]/input').click()
                     if ws['G50'].value != None : 
                        Search4 = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[2]/div/ul/li/input').click()
                        driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[2]/div/ul/li/input').send_keys(ws['G50'].value)
                        time.sleep(1)
                        driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[2]/div/ul/li/input').send_keys(Keys.ENTER)
                        
                     if ws['G51'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[2]/td[3]/input').send_keys(ws['G51'].value)
            if ws['G53'].value == "AND" or "OR" : 
                  time.sleep(1)
                  if ws['G53'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/div/a').click()
                  if ws['G53'].value == "AND" :
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[1]/div[1]/input').click()
                     if ws['G54'].value != None : 
                        Search5 = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[2]/div/ul/li/input').click()
                        driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[2]/div/ul/li/input').send_keys(ws['G54'].value)
                        time.sleep(1)
                        driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[2]/div/ul/li/input').send_keys(Keys.ENTER)

                     if ws['G55'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[3]/input').send_keys(ws['G55'].value)

                     if ws['G53'].value == "OR" :
                        driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[1]/div[2]/input').click()
                        if ws['G54'].value != None : 
                           Search6 = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[2]/div/ul/li/input').click()
                           driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[2]/div/ul/li/input').send_keys(ws['G54'].value)
                           time.sleep(1)
                           driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[2]/div/ul/li/input').send_keys(Keys.ENTER)

                        if ws['G55'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[10]/div[2]/table/tbody/tr[3]/td[3]/input').send_keys(ws['G55'].value)
         if ws['G116'].value == "YES" : 
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[1]/input').click()
            if ws['G117'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[1]/td[1]/input').send_keys(ws['G117'].value)
            if ws['G118'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[1]/td[2]/input').send_keys(ws['G118'].value)
            if ws['G119'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[1]/td[3]/input').send_keys(ws['G119'].value)
            if ws['G120'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[1]/td[4]/select').send_keys(ws['G120'].value)
         
            if ws['G122'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/div[2]/a').click()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[2]/td[1]/input').send_keys(ws['G122'].value)
            if ws['G123'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[2]/td[2]/input').send_keys(ws['G123'].value)
            if ws['G124'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[2]/td[3]/input').send_keys(ws['G124'].value)
            if ws['G125'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[2]/td[4]/select').send_keys(ws['G125'].value)
         
            if ws['G127'].value != None : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/div[2]/a').click()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[3]/td[1]/input').send_keys(ws['G127'].value)
            if ws['G128'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[3]/td[2]/input').send_keys(ws['G128'].value)
            if ws['G129'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[3]/td[3]/input').send_keys(ws['G129'].value)
            if ws['G130'].value != None : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[11]/div[2]/table/tbody/tr[3]/td[4]/select').send_keys(ws['G130'].value)

         if ws['G57'].value == "YES" : 
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[1]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[1]/div/div[2]/div/div[1]/div/div[1]/div[1]/input').click()
            if ws['G58'].value == "Calculate from joining date" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[1]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/input').click()
         
            if ws['G58'].value == "Calculate after probation period" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[1]/div/div[2]/div/div[2]/div/div/div[1]/div[2]/input').click()
         
            if ws['G59'].value == "Credit half month's leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[1]/div/div[2]/div/div[2]/div/div/div[2]/div[1]/input').click()
         
            if ws['G59'].value == "Credit full month's leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[1]/div/div[2]/div/div[2]/div/div/div[2]/div[2]/input').click()

#Credit on accrual basis
         if ws['G61'].value == "YES" : 
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[1]/div[1]/input').click()
            if ws['G62'].value == "Begin of month" or "End of month" : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[3]/div[1]/input').click()
                  if ws['G62'].value == "Begin of month" : 
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[5]/div[1]/input').click()
                  if ws['G62'].value == "End of month" :
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[5]/div[2]/input').click()
            if ws['G62'].value == "Begin of Quarter" or "End of Quarter" : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[3]/div[2]/input').click()
                  if ws['G62'].value == "Begin of Quarter" :
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[4]/div[1]/input').click()
                  if ws['G62'].value == "End of Quarter" :  
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[4]/div[2]/input').click()
            if ws['G62'].value == "Biannual" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[3]/div[3]/input').click()
            if ws['G63'].value == "YES" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[3]/div[4]/div/input').click()
               if ws['G65'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[6]/div[1]/div[3]/div/div[1]/div/input').click()
               if ws['G66'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[6]/div[1]/div[3]/div/div[2]/div/input').click()
               if ws['G67'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[6]/div[1]/div[3]/div/div[3]/div/input').click()
               if ws['G68'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[6]/div[1]/div[4]/div[1]/div/input').click()
               if ws['G69'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[6]/div[1]/div[4]/div[2]/div/input').click()
               if ws['G70'].value == "YES" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[6]/div[1]/div[4]/div[3]/div/input').click()
               if ws['G44'].value != "YES" :  
                  if ws['G71'].value == "End of the Year" : driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[6]/div[2]/div/div/div/input').click()
            if ws['G73'].value != "NA" : 
                  if ws['G71'].value != "End of the Year" :
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[3]/div[5]/div/input').click()
                     driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[2]/div/div[2]/div/div/div/div[3]/div[6]/select').send_keys(ws['G73'].value)

#Credit on Tenure basis
         if ws['G44'].value != "YES" : 
            if ws['G75'].value == "YES" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[1]/div/a/i').click()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[1]/div[1]/input').click()
               if ws['G76'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[1]/td[1]/select').send_keys(ws['G76'].value)
               if ws['G77'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[1]/td[2]/select').send_keys(ws['G77'].value)
               if ws['G78'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[1]/td[3]/input').send_keys(ws['G78'].value)
               if ws['G80'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/div[3]/a').click()
               if ws['G80'].value != None :    
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[2]/td[1]/select').send_keys(ws['G80'].value)
               if ws['G81'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[2]/td[2]/select').send_keys(ws['G81'].value)
               if ws['G82'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[2]/td[3]/input').send_keys(ws['G82'].value)
               if ws['G84'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/div[3]/a').click()
               if ws['G84'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[3]/td[1]/select').send_keys(ws['G84'].value)
               if ws['G85'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[3]/td[2]/select').send_keys(ws['G85'].value)
               if ws['G86'].value != None : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/div/table/tbody/tr[3]/td[3]/input').send_keys(ws['G86'].value)

#Allow half-day
         if ws['G112'].value == "YES" : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[4]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[4]/div/div[2]/div/div/div/div/div[1]/input').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[4]/div/div[2]/div/div/div/div/div[4]/div/input').click()

#Carry forward unused leave
         if ws['G207'].value == "YES" : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[1]/div/div[1]/div[1]/input').click()
            if ws['G208'].value == "Carry forward all unused leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[1]/div/div[3]/div[1]/input').click()
            if ws['G208'].value == "Carry forward only" :    
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[1]/div/div[3]/div[2]/label').click()
               if ws['G209'].value == "Fixed" or "Percentage" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[2]/div/div[1]/select').send_keys(ws['G209'].value)
               if ws['G210'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[2]/div/div[1]/input').clear()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[2]/div/div[1]/input').send_keys(ws['G210'].value)
               if ws['G211'].value == "Discard" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[2]/div/div[2]/div[1]/input').click()
               if ws['G211'].value == "Leave Encashment" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[2]/div/div[2]/div/div[2]/div[2]/input').click()

#Encash leave while F&F?
         if ws['G200'].value == "YES" :  
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[1]/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[1]/div[2]/div/div/div/div[1]/div[1]/input').click()
            if ws['G201'].value == "Encash all unused leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[1]/div[2]/div/div/div/div[3]/div[1]/input').click()
            if ws['G201'].value == "Encash only" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[1]/div[2]/div/div/div/div[3]/div[2]/input').click()
               if ws['G202'].value == "Fixed" or "Percentage" : 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[1]/div[2]/div/div/div/div[3]/div[3]/select').send_keys(ws['G202'].value)
               if ws['G203'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[1]/div[2]/div/div/div/div[3]/div[3]/input').clear()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[1]/div[2]/div/div/div/div[3]/div[3]/input').send_keys(ws['G203'].value)

#Count intervening Holidays/weeklys offs as leave
         if ws['G168'].value == "YES" :  
            driver.find_element_by_xpath('//*[@id="leavePolicyAccordion"]/div[5]/div/div[3]/div[2]/div[1]/div[1]/div/a/i').click()
            driver.find_element_by_xpath('//*[@id="LeavePolicy_InterveningHolidays_status_0"]').click()
            if ws['G169'].value == "YES" :
               driver.find_element_by_xpath('//*[@id="LeavePolicy_InterveningHolidays_count_intervening_holidays_1"]').click()
            if ws['G170'].value == "YES" :
               driver.find_element_by_xpath('//*[@id="LeavePolicy_InterveningHolidays_count_intervening_holidays_0"]').click()
               if ws['G171'].value == "YES" :
                  driver.find_element_by_xpath('//*[@id="LeavePolicy_InterveningHolidays_count_intervening_holidays_2"]').click()

#Can Employee apply leave prefixed or suffixed to Holidays and Weekly offs?
         if ws['G173'].value == "YES" :  
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[1]/div/a/i').click()
            driver.find_element_by_xpath('//*[@id="LeavePolicyPrefixSuffix_status_0"]').click() 
            if ws['G174'].value == "Count prefixed Weekly Offs as leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[1]/div[1]/input').click()
            if ws['G174'].value == "Block leave after Weekly Offs" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[1]/div[2]/input').click()
            if ws['G174'].value == "Allow leave after Weekly Offs" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[1]/div[3]/input').click()

            if ws['G175'].value == "Count suffixed Weekly Offs as leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[1]/div[5]/input').click()
            if ws['G175'].value == "Block leave after Weekly Offs" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[1]/div[6]/input').click()
            if ws['G175'].value == "Allow leave after Weekly Offs" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[1]/div[7]/input').click()
         
            if ws['G176'].value == "Count prefixed Holiday as leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[2]/div[1]/input').click()
            if ws['G176'].value == "Block leave after Holiday" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[2]/div[2]/input').click()
            if ws['G176'].value == "Allow leave after Holiday" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[2]/div[3]/input').click()
         
            if ws['G177'].value == "Count suffixed Holiday as leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[2]/div[5]/input').click()
            if ws['G177'].value == "Block leave after Holiday" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[2]/div[6]/input').click()
            if ws['G177'].value == "Allow leave after Holiday" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[1]/div[2]/div/div/div/div[3]/div[2]/div[7]/input').click()
            
      

#Allow Past dated Leave applications
         if ws['G94'].value == "YES" : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[2]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[2]/div/div[2]/div/div/div/div[1]/div[1]/input').click()
            if ws['G95'].value != None :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[2]/div/div[2]/div/div/div/div[3]/input').clear()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[2]/div/div[2]/div/div/div/div[3]/input').send_keys(ws['G95'].value)

#Can Employees club this leave with any other leave
         if ws['G181'].value == "YES" : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[3]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[3]/div/div[2]/div/div[1]/div/div[1]/div[1]/input').click()
            if ws['G182'].value == "Allow clubbing with any available leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[3]/div/div[2]/div/div[1]/div/div[3]/div[1]/input').click()
            if ws['G182'].value == "Allow clubbing only with following leave" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[3]/div/div[2]/div/div[1]/div/div[3]/div[2]/input').click()
               if ws['G183'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[3]/div/div[2]/div/div[2]/div/div/ul/li/input').send_keys(ws['G183'].value)
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[3]/div/div[2]/div/div[2]/div/div/ul/li/input').send_keys(Keys.ENTER)

#Can employees apply for more than their available Leave balance
         if ws['G185'].value == "YES" :  
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[1]/div[1]/input').click()
            if ws['G186'].value == "Count excess leave as Paid by default" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[3]/div[1]/input').click()
            if ws['G186'].value == "Count excess leave as Unpaid by default" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[3]/div[2]/input').click()
            if ws['G186'].value == "Utilize from" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[3]/div[3]/input').click()
            if ws['G188'].value == "YES" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[3]/div[4]/div/input[2]').click()
            if ws['G189'].value == "YES" : 
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[3]/div[5]/div/input[2]').click()
            if ws['G190'].value != None :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[3]/div[6]/div[1]/input').clear()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[2]/div/div/div/div[3]/div[6]/div[1]/input').send_keys(ws['G190'].value)

#Allow intermittent leave encashment?
         if ws['G25'].value != "Continuous Cycle" :
            if ws['G194'].value == "YES" :  
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[1]/div/a/i').click()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[1]/div[1]/input').click()
               if ws['G195'].value == "Accrued balance" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[1]/div[1]/input').click()
               if ws['G195'].value == "Carry forward balance" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[1]/div[2]/input').click()
               if ws['G195'].value == "Both Accrued & Carry forward" :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[1]/div[3]/input').click()
               if ws['G196'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/input').clear()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[2]/input').send_keys(ws['G196'].value)
               if ws['G197'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[4]/input').clear()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[4]/input').send_keys(ws['G197'].value)
               if ws['G198'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[6]/div[4]/div/input').clear()
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[3]/div/div[2]/div/div/div/div[3]/div[6]/div[4]/div/input').send_keys(ws['G198'].value)

#Allow date specific leave
         if ws['G215'].value == "YES" :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[1]/div[1]/input').click()
            if ws['G216'].value != None :
               SearchClick10 = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[1]/div/ul/li/input').click()
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[1]/div/ul/li/input').send_keys(ws['G216'].value)
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[1]/div/ul/li/input').send_keys(Keys.ENTER)
            if ws['G217'].value == "YES" :
               driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[2]/div/input[2]').click()
               if ws['G218'].value != None :
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[4]/div/table/tbody/tr[1]/td[1]/input').send_keys(ws['G218'].value)
               if ws['G219'].value != None :
                  Search7 = driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[4]/div/table/tbody/tr[1]/td[2]/input').click() 
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[4]/div/table/tbody/tr[1]/td[2]/input').send_keys(ws['G219'].value)
                  driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[4]/div/div[2]/div/div/div/div[3]/div[4]/div/table/tbody/tr[1]/td[2]/input').send_keys(Keys.ENTER)
                                            
#Does the usability of this leave balance depend upon another leave?
         if ws['G223'].value == "YES" :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[5]/div/div[1]/div/a/i').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[4]/div[3]/div/div[5]/div/div[3]/div[2]/div[2]/div[4]/div/div[5]/div/div[2]/div/div/div/div/div[1]/input').click()
         # To save Leave Policy
         driver.execute_script("window.scrollTo(0, -500)") 
         driver.find_element_by_xpath('/html/body/div[2]/div/section/form/div/div[1]/div/div[2]/div/input').click()
         time.sleep(3)

#Leave Settings and Unpaid Leave Reason
def Leave_Settings():
   global driver
   wb=Workbook()
   wb_path= r"D:\Leaves Process Mapping_v3.xlsx"
   wbk = openpyxl.load_workbook(wb_path,data_only=True)

   a = ["5. Leave Settings"]
 
   for i in a :
      ws = wbk[i]
      driver.get(WebLink.get() + '/settings/leaves/settings')
      if ws['G7'].value != None :
         driver.find_element_by_xpath('//*[@id="TenantLeavesSettings_unpaid_refresh"]').send_keys(ws['G7'].value)
      if ws['G26'].value != None :
         driver.find_element_by_xpath('//*[@id="employee-search-tokenfield"]').send_keys(ws['G26'].value)
      if ws['G52'].value != None :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[3]/div/input').clear()
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[3]/div/input').send_keys(ws['G52'].value)
      if ws['G40'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[4]/div/input').click()
         time.sleep(1)
      if ws['G41'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[5]/div/input').click()
         time.sleep(1)
      if ws['G36'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[6]/div/input').click()
         time.sleep(1)
      if ws['G37'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[7]/div/input').click()
         time.sleep(1)
      if ws['G38'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[8]/div/input').click()
         time.sleep(1)
      if ws['G8'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[9]/div/input').click()
         time.sleep(1)
      if ws['G9'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[10]/div/input').click()
         time.sleep(1)
      if ws['G11'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[13]/div/input').click()
         time.sleep(2)
      if ws['G10'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[14]/div/input').click()
         time.sleep(1)
      if ws['G12'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[15]/div/input').click()
         time.sleep(1)
         # if ws['G13'].value =="NO" :     
         #    driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[16]/div[1]/div/input').click()
         #    time.sleep(1)
         # if ws['G14'].value =="NO" :     
         #    driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[16]/div[2]/div/input').click()
         #    time.sleep(1)
         # if ws['G15'].value =="NO" :     
         #    driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[16]/div[3]/div/input').click()
         #    time.sleep(1)
      if ws['G6'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[17]/div/input').click()
         time.sleep(1)
      if ws['G45'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[18]/div/input').click()
         time.sleep(1)
      if ws['G46'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[19]/div/input').click()
         time.sleep(1)
      if ws['G56'].value =="YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[6]/div[21]/div/input').click()
         time.sleep(1)
      driver.execute_script("window.scrollTo(0, 0)") 
      if ws['G17'].value =="During fixed duration" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[1]/div/div[1]/div/input').click()
         time.sleep(1)
         if ws['G19'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[2]/div/div/input[1]').click()
            time.sleep(1)
            Y1=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').click()
            time.sleep(1)
            YY1=ws['N19'].value
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(YY1)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(Keys.ENTER)
            time.sleep(1)
            M1=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').click()
            time.sleep(1)
            MM1=ws['O19'].value
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(MM1)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(Keys.ENTER)
            time.sleep(1)
            
            alldates1=driver.find_elements_by_xpath('//*[@id="ui-datepicker-div"]/table/tbody/tr/td/a')

            time.sleep(1)
            for dateelement1 in alldates1 :
               date1 = dateelement1.text
               DD1=ws['P19'].value
               if date1 == DD1 :
                  dateelement1.click()
                  

         if ws['G20'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[2]/div/div/input[2]').click()
            time.sleep(1)
            Y2=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(ws['N20'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(Keys.ENTER)
            time.sleep(1)
            M2=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(ws['O20'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(Keys.ENTER)
            time.sleep(1)
            
            alldates2=driver.find_elements_by_xpath('//*[@id="ui-datepicker-div"]/table/tbody/tr/td/a')

            time.sleep(1)
            for dateelement2 in alldates2 :
               date2 = dateelement2.text

               if date2 == ws['P20'].value :
                  dateelement2.click()
                  

      driver.execute_script("window.scrollTo(0, 0)") 
      time.sleep(1)
      if ws['G17'].value =="During defined days from cycle end date" : 
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[1]/div/div[2]/div/input').click()
         time.sleep(1)
         if ws['G22'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[3]/div/div/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[3]/div/div/input').send_keys(ws['G22'].value)
            time.sleep(1)
      if ws['G58'].value != None :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[5]/div/div/ul/li/input').clear()
         time.sleep(1)
         Leaves_NOT_to_be_counted=driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[5]/div/div/ul/li/input').click()
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[5]/div/div/ul/li/input').send_keys(ws['G58'].value)
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[7]/div[5]/div/div/ul/li/input').send_keys(Keys.ENTER)
         time.sleep(1)
      
      if ws['G17'].value =="During fixed duration" :
         # Need to add date selection code
         if ws['G49'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[8]/div[1]/div/div/input[1]').click()
            time.sleep(1)
            Y3=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(ws['N49'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(Keys.ENTER)
            time.sleep(1)
            M3=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(ws['O49'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(Keys.ENTER)
            time.sleep(1)
            
            alldates3=driver.find_elements_by_xpath('//*[@id="ui-datepicker-div"]/table/tbody/tr/td/a')

            time.sleep(1)
            for dateelement3 in alldates3 :
               date3 = dateelement3.text

               if date3 == ws['P49'].value :
                  dateelement3.click()
                  

         if ws['G50'].value != None : 
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[8]/div[1]/div/div/input[2]').click()
            time.sleep(1)
            Y4=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(ws['N50'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[2]').send_keys(Keys.ENTER)
            time.sleep(1)
            M4=driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(ws['O50'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[5]/div/div/select[1]').send_keys(Keys.ENTER)
            time.sleep(1)
            
            alldates4=driver.find_elements_by_xpath('//*[@id="ui-datepicker-div"]/table/tbody/tr/td/a')

            time.sleep(1)
            for dateelement4 in alldates4 :
               date4 = dateelement4.text

               if date4 == ws['P50'].value :
                  dateelement4.click()

      time.sleep(1)
      if ws['G28'].value != None :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[8]/div[2]/select').send_keys(ws['G28'].value)
      time.sleep(1)
      if ws['G30'].value == "YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[9]/div[1]/div/input').click()
         time.sleep(1)
         if ws['G31'].value!=None :
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[9]/div[2]/div[1]/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[9]/div[2]/div[1]/input').send_keys(ws['G31'].value)
         if ws['G32'].value != None :
            time.sleep(1)
            Escalate_To = driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[9]/div[2]/div[2]/labe/div/input').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[9]/div[2]/div[2]/labe/div/input').send_keys(ws['G32'].value)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[9]/div[2]/div[2]/labe/div/input').send_keys(Keys.ENTER)
         time.sleep(1)
         driver.execute_script("window.scrollTo(0, 0)")
         time.sleep(1)
         driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div/form/div[2]/div/input').click()
      time.sleep(2)

# Unpaid Leave Reason:
def Unpaid_Leave_Reason():
   global driver
   wb=Workbook()
   wb_path= r"D:\Leaves Process Mapping_v3.xlsx"
   wbk = openpyxl.load_workbook(wb_path,data_only=True)

   a = ["5. Leave Settings"]
 
   for i in a :

      ws = wbk[i]
      driver.get(WebLink.get() + '/settings/leaves/unpaidreasons')
      if ws['G62'].value != None :
         if ws['M62'].value != None :
            time.sleep(2)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[5]/div[1]/div/input').send_keys(ws['M62'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[2]/div/input').click()
            time.sleep(2)

         if ws['N62'].value != None :
            time.sleep(2)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[5]/div[1]/div/input').send_keys(ws['N62'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[2]/div/input').click()
            time.sleep(2)

         if ws['O62'].value != None :
            time.sleep(2)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[5]/div[1]/div/input').send_keys(ws['O62'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[2]/div/input').click()
            time.sleep(2)

         if ws['P62'].value != None :
            time.sleep(2)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[5]/div[1]/div/input').send_keys(ws['P62'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[2]/div/input').click()
            time.sleep(2)

         if ws['Q62'].value != None :
            time.sleep(2)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[5]/div[1]/div/input').send_keys(ws['Q62'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[2]/div/input').click()
            time.sleep(2)

         if ws['R62'].value != None :
            time.sleep(2)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[4]/div/a').click()
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[5]/div[1]/div/input').send_keys(ws['R62'].value)
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/section/div/div/div[1]/form/div[2]/div/input').click()
            time.sleep(2)

# Unpaid Leave
def Unpaid_Leave():
   global driver
   wb=Workbook()
   wb_path= r"D:\Leaves Process Mapping_v3.xlsx"
   wbk = openpyxl.load_workbook(wb_path)

   a = ["4. LOP or Unpaid"]
 
   for i in a :
      ws = wbk[i]
      driver.get(WebLink.get() + '/settings/leaves/unpaid')
      if ws['G6'].value != None :
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[1]/div/input').clear()
         time.sleep(1)
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[1]/div/input').send_keys(ws['G6'].value)
         time.sleep(2)
      driver.execute_script("window.scrollTo(0, 200)")
      if ws['G16'].value == "YES" :
         time.sleep(1)
         if ws['G17'].value == "YES" :
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[4]/div/div[1]/input[2]').click()
            time.sleep(1)
         if ws['G18'].value == "YES" :
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[4]/div/div[2]/input[2]').click()
            time.sleep(2)
            if ws['G19'].value == "YES" :
               driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[4]/div/div[4]/input').click()
               time.sleep(1)
      
      if ws['G13'].value == "YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[5]/div/div[1]/input').click()
         time.sleep(1)
         if ws['G14'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[6]/div/div/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[6]/div/div/input').send_keys(ws['G14'].value)
      time.sleep(1)
      if ws['G24'].value != None :
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[7]/div/div/input').clear()
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[7]/div/div/input').send_keys(ws['G24'].value)
            time.sleep(1)
      if ws['G11'].value == "YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[8]/div/div[1]/input').click()
         time.sleep(1)
      if ws['G26'].value == "YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[9]/div/input[2]').click()
         time.sleep(1)
      if ws['G28'].value == "YES" :
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[10]/div/input').click()
         time.sleep(1)
      if ws['G8'].value == "AND" :
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[11]/div[1]/input').click()
         if ws['G9'].value != None :
            Restrictions=driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[12]/div/div/ul/li/input').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[12]/div/div/ul/li/input').send_keys(ws['G9'].value)
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[12]/div/div/ul/li/input').send_keys(Keys.ENTER)
            time.sleep(1)
      if ws['G8'].value == "OR" :
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[11]/div[2]/input').click()
         if ws['G9'].value != None :
            Restrictions1=driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[12]/div/div/ul/li/input').click()
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[12]/div/div/ul/li/input').send_keys(ws['G9'].value)
            driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[12]/div/div/ul/li/input').send_keys(Keys.ENTER)
            time.sleep(1)
      if ws['G22'].value != None :
         ApprovalFlow = driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[13]/select').click()
         driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[8]/div/div[13]/select').send_keys(ws['G9'].value)
         time.sleep(1)
      driver.execute_script("window.scrollTo(0, 0)")
      time.sleep(1)
      driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/section/div/div/form/div[4]/div/input').click()

time.sleep(2)

# --------------------------------------------------------------------------------------------------------------
# ---------------------------------------------- Tkinter -------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------

InfoForTHeUser ="Please make sure that you have used 'Leaves Process Mapping_v3' to take the data and it is present in 'D:' Drive."

driver = None

root  = tk.Tk()

img = PhotoImage(file=resource_path_logo('/Img/Dbox4.png'))

Label(root,image=img,height=60).grid(sticky=NE,padx=20,pady=20)

root.title('DarwinBox')
#width then hight
root.geometry('600x520+50+50')

root['bg'] = '#5252ff'

tk.Label(root,text="Client Instance / Website link ->").grid(row=3,column=1,padx=5,pady=5)
WebLink = StringVar()
name1 = tk.Entry(root, textvariable=WebLink,width=30)
name1.grid(row=3,column=2,padx=5,pady=5)

tk.Label(root,text="User ID / Email ID ->",activebackground='white').grid(row=4,column=1,padx=5,pady=5)
username1 = StringVar()
name2 = tk.Entry(root, textvariable=username1,width=30)
name2.grid(row=4,column=2,padx=5,pady=5)

tk.Label(root,text="Password ->").grid(row=5,column=1,padx=5,pady=5)
password1 = StringVar()
name3 = tk.Entry(root, textvariable=password1,show="*",width=30)
name3.grid(row=5,column=2,padx=5,pady=5)

b7 = tk.Label(root,text="*** Please provide all the inputs ***",width=55).grid(row=8,column=1,padx=5,pady=5,columnspan=4)

b6 = tk.Label(root,text="",background='#5252ff').grid(row=11,column=1,padx=10,pady=1,columnspan=4)
 
b1 = tk.Button(root, text='Chrome Open', command=on_open,width=40,relief=RAISED,activebackground='Grey').grid(row=12,column=1,padx=5,pady=5,columnspan=4)

b2 = tk.Button(root, text='Login & Admin', command=UserLogin,width=40,relief=RAISED,activebackground='Grey').grid(row=13,column=1,padx=5,pady=5,columnspan=4)

b3 = tk.Button(root, text='Add Leave Policy', command=Add_Leave_Policy,width=40,relief=RAISED,activebackground='Grey').grid(row=14,column=1,padx=5,pady=5,columnspan=4)

b4 = tk.Button(root, text='Leave Settings', command=Leave_Settings,width=40,relief=RAISED,activebackground='Grey').grid(row=15,column=1,padx=5,pady=5,columnspan=4)

b9 = tk.Button(root, text='Unpaid Leave Reason', command=Unpaid_Leave_Reason,width=40,relief=RAISED,activebackground='Grey').grid(row=16,column=1,padx=5,pady=5,columnspan=4)

b8 = tk.Button(root, text='Unpaid Leave', command=Unpaid_Leave,width=40,relief=RAISED,activebackground='Grey').grid(row=17,column=1,padx=5,pady=5,columnspan=4)

b5 = tk.Button(root, text='Chrome Close', command=on_close,width=40,relief=RAISED,activebackground='Grey').grid(row=18,column=1,padx=10,pady=5,columnspan=4)

root.mainloop()
